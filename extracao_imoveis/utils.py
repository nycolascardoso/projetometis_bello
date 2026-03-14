import openpyxl
from datetime import datetime
from extracao_imoveis.config import PLANILHA_PATH

def carregar_planilha(caminho):
    """Carrega a planilha Excel a partir do caminho informado."""
    try:
        return openpyxl.load_workbook(caminho)
    except Exception as e:
        print(f"❌ Erro ao carregar a planilha: {e}")
        raise

def atualizar_status_consultar(aba, cnpj_cpf, status):
    """
    Atualiza o status e a última atualização na aba 'Consultar'.
    
    Estrutura esperada da aba "Consultar":
      - Coluna 1: CNPJ/CPF
      - Coluna 2: Status
      - Coluna 3: Última Atualização
      - Coluna 4: Tipo de Doc.
      - Coluna 5: Processar
    
    Se o status informado não indicar erro (ou seja, não contiver "Erro"),
    a coluna "Processar" (coluna 5) é redefinida para "Não".
    """
    try:
        for row in aba.iter_rows(min_row=2, values_only=False):
            if row[0].value == cnpj_cpf:
                row[1].value = status  # Atualiza o status
                row[2].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Atualiza a data/hora
                if status == "Finalizado":
                    row[4].value = "Não"  # Define "Processar" como "Não" só quando concluído com sucesso
                print(f"🔄 Status atualizado para '{status}' no CNPJ/CPF {cnpj_cpf}.")
                return
        print(f"⚠️ CNPJ/CPF {cnpj_cpf} não encontrado na aba 'Consultar'.")
    except Exception as e:
        print(f"❌ Erro ao atualizar status: {e}")
        raise

def salvar_dados_na_aba(aba, dados):
    """
    Adiciona uma nova linha com os dados na aba fornecida.

    Parâmetros:
      - aba: aba onde os dados serão adicionados.
      - dados: lista com os valores a serem inseridos.
    """
    try:
        aba.append(dados)
        print(f"✅ Dados salvos na aba '{aba.title}': {dados}")
    except Exception as e:
        print(f"❌ Erro ao salvar dados na aba '{aba.title}': {e}")
        raise

def salvar_dados_reprocessamento(aba, dados, caminho):
    """
    Adiciona uma nova linha com os dados na aba 'Reprocessamento' e salva a planilha.
    Essa função é útil para salvar dados oriundos de linhas marcadas para reprocessamento.
    
    Parâmetros:
      - aba: a aba "Reprocessamento" da planilha.
      - dados: lista com os valores a serem adicionados.
      - caminho: caminho da planilha para garantir que os dados sejam salvos corretamente.
    """
    try:
        aba.append(dados)
        aba.parent.save(caminho)
        print(f"🔁 Dados salvos na aba '{aba.title}' (Reprocessamento): {dados}")
    except Exception as e:
        print(f"❌ Erro ao salvar dados na aba '{aba.title}' (Reprocessamento): {e}")
        raise

def obter_linhas_para_processamento(aba):
    """
    Retorna uma lista de linhas da aba "Consultar" que devem ser processadas,
    ou seja, aquelas em que a coluna "Processar" (coluna 5) está marcada como "Sim".

    Retorna:
      Uma lista de dicionários contendo os dados essenciais (CNPJ/CPF e Tipo de Doc.)
      para cada linha marcada como "Sim".
    """
    linhas_processar = []
    try:
        for row in aba.iter_rows(min_row=2, values_only=True):
            cnpj_cpf = row[0]
            tipo_doc = row[3]
            processar = row[4]

            # Se tipo_doc é uma fórmula (não avaliada sem data_only), infere pelo tamanho do documento
            if not tipo_doc or str(tipo_doc).strip().startswith('='):
                tipo_doc = "CNPJ proprietário" if len(str(cnpj_cpf)) == 18 else "CPF proprietário"

            if processar and str(processar).strip().lower() == "sim":
                linhas_processar.append({"cnpj_cpf": cnpj_cpf, "tipo_doc": tipo_doc})
        print(f"📌 {len(linhas_processar)} linhas encontradas para processamento.")
        return linhas_processar
    except Exception as e:
        print(f"❌ Erro ao obter linhas para processamento: {e}")
        raise
