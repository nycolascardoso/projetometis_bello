import openpyxl
from datetime import datetime
from config import PLANILHA_PATH

def carregar_planilha(caminho):
    """Carrega a planilha Excel a partir do caminho informado."""
    try:
        return openpyxl.load_workbook(caminho)
    except Exception as e:
        print(f"Erro ao carregar a planilha: {e}")
        raise

def atualizar_status_consultar(aba, cnpj_cpf, status, caminho):
    """
    Atualiza o status e a última atualização na aba 'Consultar'.
    
    A estrutura esperada da aba "Consultar" é:
      Coluna 1: CNPJ/CPF
      Coluna 2: Status
      Coluna 3: Última Atualização
      Coluna 4: Tipo de Doc.
      Coluna 5: Processar

    Se o status informado não indicar erro (ou seja, não contiver "Erro"),
    a coluna "Processar" (coluna 5) é redefinida para "Não".
    
    Parâmetros:
      - aba: objeto da aba "Consultar" da planilha.
      - cnpj_cpf: CNPJ/CPF cujo status será atualizado.
      - status: novo status (ex.: "Em progresso", "Finalizado", "Erro: ...").
      - caminho: caminho para salvar a planilha após a atualização.
    """
    try:
        for row in aba.iter_rows(min_row=2, values_only=False):
            if row[0].value == cnpj_cpf:
                row[1].value = status  # Atualiza o status
                row[2].value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')  # Atualiza a data/hora
                # Se o status não for de erro, redefinir a coluna "Processar" para "Não"
                if "Erro" not in status:
                    row[4].value = "Não"
                aba.parent.save(caminho)
                print(f"Status atualizado para '{status}' no CNPJ/CPF {cnpj_cpf}.")
                break
    except Exception as e:
        print(f"Erro ao atualizar status: {e}")
        raise

def salvar_dados_na_aba(aba, dados):
    """
    Adiciona uma nova linha com os dados na aba fornecida e salva a planilha.

    Parâmetros:
      - aba: a aba onde os dados serão adicionados.
      - dados: lista com os valores a serem inseridos como nova linha.
    """
    try:
        aba.append(dados)
        aba.parent.save(PLANILHA_PATH)
        print(f"Dados salvos na aba '{aba.title}': {dados}")
    except Exception as e:
        print(f"Erro ao salvar dados na aba '{aba.title}': {e}")
        raise

def salvar_dados_reprocessamento(aba, dados):
    """
    Adiciona uma nova linha com os dados na aba 'Reprocessamento' e salva a planilha.
    Essa função é útil para salvar dados oriundos de linhas marcadas para reprocessamento.
    
    Parâmetros:
      - aba: a aba "Reprocessamento" da planilha.
      - dados: lista com os valores a serem adicionados.
    """
    try:
        aba.append(dados)
        aba.parent.save(PLANILHA_PATH)
        print(f"Dados salvos na aba '{aba.title}' (Reprocessamento): {dados}")
    except Exception as e:
        print(f"Erro ao salvar dados na aba '{aba.title}' (Reprocessamento): {e}")
        raise

def obter_linhas_para_processamento(aba):
    """
    Retorna uma lista de linhas da aba "Consultar" que devem ser processadas,
    ou seja, aquelas em que a coluna "Processar" (coluna 5) está marcada como "Sim".

    Considera a seguinte estrutura:
      Coluna 1: CNPJ/CPF
      Coluna 2: Status
      Coluna 3: Última Atualização
      Coluna 4: Tipo de Doc.
      Coluna 5: Processar

    Retorna:
      Uma lista de dicionários contendo os dados essenciais (CNPJ/CPF e Tipo de Doc.)
      para cada linha marcada como "Sim".
    """
    linhas_processar = []
    try:
        # Itera a partir da segunda linha (ignorando o cabeçalho)
        for row in aba.iter_rows(min_row=2, values_only=True):
            cnpj_cpf = row[0]
            tipo_doc = row[3]
            processar = row[4]
            if processar and processar.strip().lower() == "sim":
                linhas_processar.append({
                    "cnpj_cpf": cnpj_cpf,
                    "tipo_doc": tipo_doc
                })
        print(f"Foram encontradas {len(linhas_processar)} linhas para processamento.")
        return linhas_processar
    except Exception as e:
        print(f"Erro ao obter linhas para processamento: {e}")
        raise
