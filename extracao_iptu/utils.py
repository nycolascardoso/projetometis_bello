import openpyxl
import os
from datetime import datetime
from extracao_iptu.config import PLANILHA_PATH


def carregar_planilha():
    """Carrega a planilha Excel garantindo que não esteja corrompida."""
    try:
        if not os.path.exists(PLANILHA_PATH):
            print(f"❌ Arquivo da planilha não encontrado: {PLANILHA_PATH}")
            return None
        planilha = openpyxl.load_workbook(PLANILHA_PATH)
        print("📂 Planilha carregada com sucesso.")
        return planilha
    except Exception as e:
        print(f"❌ Erro ao carregar a planilha: {e}")
        return None


def salvar_planilha(planilha):
    """Salva a planilha via arquivo temporário para evitar corrupção em caso de falha de I/O."""
    try:
        if planilha is None:
            print("⚠️ Planilha não está carregada. Tentando reabrir...")
            planilha = carregar_planilha()

        temp_path = PLANILHA_PATH.replace(".xlsx", "_temp.xlsx")

        try:
            planilha.save(temp_path)
        except ValueError:
            print("⚠️ Planilha fechada inesperadamente. Reabrindo antes de salvar...")
            planilha = carregar_planilha()
            planilha.save(temp_path)

        os.replace(temp_path, PLANILHA_PATH)

    except Exception as e:
        print(f"❌ Erro ao salvar a planilha: {e}")


def atualizar_status_iptu(aba, codigo_imovel, status):
    """Atualiza a coluna 'IPTU Extraído?' (L) para o imóvel indicado. Não salva — salve após chamar."""
    try:
        for row in aba.iter_rows(min_row=2, values_only=False):
            if str(row[1].value) == str(codigo_imovel):
                row[11].value = status
                return True
        print(f"⚠️ Código do imóvel {codigo_imovel} não encontrado na aba 'Link de Imóveis'.")
        return False
    except Exception as e:
        print(f"❌ Erro ao atualizar status IPTU [{codigo_imovel}]: {e}")
        return False


def salvar_dados_na_aba(aba, dados):
    """Adiciona linhas na aba indicada. Não salva — salve após chamar."""
    try:
        if not aba:
            print("❌ Aba inválida ao tentar salvar dados.")
            return
        for linha in dados:
            aba.append(linha)
    except Exception as e:
        print(f"❌ Erro ao salvar dados na aba '{aba.title}': {e}")


def atualizar_dados_imovel(aba, codigo_imovel, localizacao, tipologia, estrutura, utilizacao, proprietario):
    """Atualiza colunas M–Q na aba 'Link de Imóveis'. Não salva — salve após chamar."""
    try:
        for row in aba.iter_rows(min_row=2, values_only=False):
            if str(row[1].value) == str(codigo_imovel):
                if utilizacao and "baldio" in str(utilizacao).lower():
                    localizacao = tipologia = estrutura = utilizacao = proprietario = "Baldio s/uso"
                row[12].value = localizacao
                row[13].value = tipologia
                row[14].value = estrutura
                row[15].value = utilizacao
                row[16].value = proprietario
                return True
        print(f"⚠️ Código do imóvel {codigo_imovel} não encontrado na aba 'Link de Imóveis'.")
        return False
    except Exception as e:
        print(f"❌ Erro ao atualizar dados do imóvel {codigo_imovel}: {e}")
        return False


def obter_linhas_para_processamento(aba):
    """Retorna imóveis da aba 'Consultar' onde 'Processar' == 'Sim'."""
    linhas = []
    try:
        for row in aba.iter_rows(min_row=2, values_only=True):
            codigo_imovel = row[0]
            processar = row[4]
            if processar and str(processar).strip().lower() == "sim":
                linhas.append({"codigo_imovel": codigo_imovel})
        print(f"🔍 {len(linhas)} imóveis identificados para processamento.")
        return linhas
    except Exception as e:
        print(f"❌ Erro ao obter linhas para processamento: {e}")
