"""
Funções utilitárias para o app Streamlit do Metis.
Leitura/escrita do Excel e controle do runtime de extração.
"""
import os
import json
import openpyxl
import pandas as pd
from io import BytesIO

# Caminhos base
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PLANILHA_PATH = os.path.join(BASE_DIR, 'Banco_de_Imoveis.xlsx')
STATUS_PATH = os.path.join(BASE_DIR, 'runtime_status.json')


# ---------------------------------------------------------------------------
# Leitura do Excel
# ---------------------------------------------------------------------------

def ler_consultar() -> pd.DataFrame:
    """Lê a aba 'Consultar' e retorna um DataFrame com CNPJ/CPFs cadastrados."""
    try:
        df = pd.read_excel(
            PLANILHA_PATH,
            sheet_name='Consultar',
            header=0,
            usecols='A:G',
            dtype=str
        )
        # Normaliza nomes de coluna
        df.columns = [
            'CNPJ/CPF', 'Status', 'Última Atualização',
            'Tipo de Doc.', 'Processar',
            'Relatório Prefeitura', 'Nome Relatório'
        ]
        df = df[df['CNPJ/CPF'].notna() & (df['CNPJ/CPF'].str.strip() != '')]
        return df.reset_index(drop=True)
    except Exception as e:
        return pd.DataFrame(columns=[
            'CNPJ/CPF', 'Status', 'Última Atualização',
            'Tipo de Doc.', 'Processar',
            'Relatório Prefeitura', 'Nome Relatório'
        ])


def ler_link_imoveis() -> pd.DataFrame:
    """Lê a aba 'Link de Imóveis' e retorna um DataFrame."""
    try:
        df = pd.read_excel(
            PLANILHA_PATH,
            sheet_name='Link de Imóveis',
            header=0,
            dtype=str
        )
        col_names = [
            'CNPJ/CPF', 'Código do Imóvel', 'Inscrição Imobiliária',
            'Logradouro', 'Complemento', 'Bairro', 'Situação',
            'URL', 'Status', 'Última Atualização', 'Extrair IPTU?',
            'IPTU Extraído?', 'Localização', 'Tipologia', 'Estrutura',
            'Utilização', 'Proprietário'
        ]
        # Aplica apenas até o número de colunas existentes
        df.columns = col_names[:len(df.columns)] + list(df.columns[len(col_names):])
        df = df[df['CNPJ/CPF'].notna() & (df['CNPJ/CPF'].str.strip() != '')]
        return df.reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def ler_banco_dados() -> pd.DataFrame:
    """Lê a aba 'Banco de Dados' e retorna um DataFrame."""
    try:
        df = pd.read_excel(
            PLANILHA_PATH,
            sheet_name='Banco de Dados',
            header=0,
            dtype=str
        )
        return df.reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


# ---------------------------------------------------------------------------
# Escrita no Excel
# ---------------------------------------------------------------------------

def marcar_para_processar(cnpjs_selecionados: list):
    """
    Na aba 'Consultar', define col E='Sim' apenas para os CNPJs da lista.
    Demais linhas que estavam como 'Sim' são resetadas para 'Não'.
    """
    wb = openpyxl.load_workbook(PLANILHA_PATH)
    aba = wb['Consultar']
    for row in aba.iter_rows(min_row=2, values_only=False):
        cnpj = str(row[0].value).strip() if row[0].value else ''
        if cnpj in [str(c).strip() for c in cnpjs_selecionados]:
            row[4].value = 'Sim'
        elif row[4].value and str(row[4].value).strip().lower() == 'sim':
            row[4].value = 'Não'
    wb.save(PLANILHA_PATH)


def marcar_para_extrair_iptu(codigos_selecionados: list):
    """
    Na aba 'Link de Imóveis', define col K='Sim' apenas para os códigos da lista.
    Outros imóveis que estavam como 'Sim' e já foram processados não são alterados.
    """
    wb = openpyxl.load_workbook(PLANILHA_PATH)
    aba = wb['Link de Imóveis']
    codigos_str = [str(c).strip() for c in codigos_selecionados]
    for row in aba.iter_rows(min_row=2, values_only=False):
        codigo = str(row[1].value).strip() if row[1].value else ''
        if codigo in codigos_str:
            row[10].value = 'Sim'
    wb.save(PLANILHA_PATH)


# ---------------------------------------------------------------------------
# Status do runtime
# ---------------------------------------------------------------------------

def ler_status_runtime() -> dict:
    """Lê runtime_status.json e retorna o status atual da extração."""
    try:
        if os.path.exists(STATUS_PATH):
            with open(STATUS_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
    except Exception:
        pass
    return {
        "running": False,
        "module": None,
        "current": 0,
        "total": 0,
        "current_item": "",
        "log": []
    }


def extracao_rodando() -> bool:
    """Retorna True se uma extração está em andamento."""
    return ler_status_runtime().get("running", False)


# ---------------------------------------------------------------------------
# Exportação
# ---------------------------------------------------------------------------

def gerar_export_excel(df: pd.DataFrame) -> bytes:
    """Gera bytes de um arquivo Excel (.xlsx) em memória para download."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Dados')
    return output.getvalue()


def gerar_export_csv_importacao(df_banco: pd.DataFrame) -> bytes:
    """
    Gera CSV no formato de importação da imobiliária.
    Mapeamento das colunas do 'Banco de Dados':
      - Col B (idx 1) = Inscrição Imobiliária
      - Col C (idx 2) = Ano
      - Col D (idx 3) = Descrição
      - Col E (idx 4) = Nº parc.
      - Col G (idx 6) = Dt. vcto. → vencimento_da_primeira_parcela
      - Col H (idx 7) = Vlr. parcela → valor_da_primeira_parcela
      - Col I (idx 8) = Desconto → taxa_de_desconto_%
      - Col N (idx 13) = Código Imóvel → codigo_do_encargo
    """
    cols = df_banco.columns.tolist()

    def safe_col(idx, default=''):
        return df_banco.iloc[:, idx] if idx < len(cols) else default

    export = pd.DataFrame()
    export['categoria_do_encargo'] = '1'
    export['tipo_do_encargo'] = '1'
    # Código do imóvel (col N, índice 13)
    export['codigo_do_encargo'] = safe_col(13, '')
    export['valor_total_do_carne'] = safe_col(7, '')
    export['parcelas'] = safe_col(4, '')
    export['taxa_de_desconto_%'] = safe_col(8, '')
    export['vencimento_da_primeira_parcela'] = safe_col(6, '')
    export['valor_da_primeira_parcela'] = safe_col(7, '')
    export['codigo_de_barras'] = ''
    # Complemento: "IPTU " + ano
    ano_col = safe_col(2, '2025')
    export['complemento_despesa'] = 'IPTU ' + ano_col.astype(str)
    export['calcular_taxa_de_administracao_sobre_o_servico'] = '0'
    export['conta_bancaria'] = '6'

    csv_bytes = export.to_csv(index=False, sep=';', encoding='utf-8-sig').encode('utf-8-sig')
    return csv_bytes
